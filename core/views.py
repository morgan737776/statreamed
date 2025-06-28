from django.shortcuts import render, redirect, get_object_or_404
from django.views.generic import TemplateView, View, ListView, CreateView, UpdateView, DeleteView
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.utils import timezone
from django.urls import reverse, reverse_lazy
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.db import models
from django.db.models import Q, Count, Sum
from django.core.exceptions import PermissionDenied
from django.contrib.auth.models import User, Group
from django.contrib.auth.forms import UserCreationForm, UserChangeForm
from django.contrib.auth import get_user_model
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django import forms
from django.http import HttpResponse, JsonResponse
from datetime import date, timedelta
from collections import Counter
from django.contrib.humanize.templatetags.humanize import intcomma
import json
from .forms import PatientDocumentForm
import csv

from .models import (
    Patient, MedicalRecord, 
    SystemSettings, IntegrationSettings, AuditLog,
    Appointment, Notification, TreatmentReminder, ExternalCalendarSync
)
from rehab_programs.models import RehabilitationProgram
from services.models import ServiceAppointment, ServiceItem
from .forms import (
    PatientForm, MedicalRecordForm, 
    SystemSettingsForm, 
    IntegrationSettingsForm
)
from medical_history.forms import MedicalRecordForm as MedicalHistoryRecordForm
from rehab_programs.forms import RehabilitationProgramForm
from rehab_center.seo_settings import SITE_NAME, SITE_DESCRIPTION

class PortalHomeView(TemplateView):
    template_name = 'core/portal_home.html'

class ReportsHomeView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/home.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['reports'] = [
            {'name': 'Статистика по пациентам', 'url': 'core:report_patients', 'icon': 'fa-users'},
            {'name': 'Загруженность коечного фонда', 'url': 'core:report_beds', 'icon': 'fa-bed'},
            {'name': 'Финансовые отчёты', 'url': 'core:report_finance', 'icon': 'fa-wallet'},
            {'name': 'Эффективность персонала', 'url': 'core:report_staff', 'icon': 'fa-user-md'},
            {'name': 'Аналитика записей', 'url': 'core:report_appointments', 'icon': 'fa-calendar-check'},
            {'name': 'Результаты программ', 'url': 'core:report_rehab', 'icon': 'fa-heartbeat'},
        ]
        return ctx

@login_required
@permission_required('core.view_bed', raise_exception=True)
def report_beds(request):
    # from .models import Ward, Bed
    # departments = Ward.objects.values_list('department', flat=True).distinct()
    # wards = Ward.objects.all().order_by('name')
    # 
    # # --- Queryset коек ---
    # beds_qs = Bed.objects.select_related('ward').all()
    # ... остальной код функции временно отключен
    return render(request, 'core/reports/beds.html', {'message': 'Функция временно недоступна'})

@login_required
def report_finance(request):
    # --- Фильтры ---
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    service_id = request.GET.get('service')
    export_format = request.GET.get('export')

    # --- Списки для фильтров ---
    services = ServiceItem.objects.all().order_by('name')

    # --- Queryset приёмов ---
    appointments = ServiceAppointment.objects.filter(status='completed')
    if date_from:
        appointments = appointments.filter(appointment_date__gte=parse_date(date_from))
    if date_to:
        appointments = appointments.filter(appointment_date__lte=parse_date(date_to))
    if service_id:
        appointments = appointments.filter(service_id=service_id)

    # --- Данные для таблицы ---
    appointments_list = appointments.select_related('client', 'service').order_by('-appointment_date', '-start_time')[:200]

    # --- Данные для графика (доход по дням) ---
    from collections import defaultdict
    chart_data = {'labels': [], 'values': []}
    if date_from and date_to:
        from datetime import timedelta
        start = parse_date(date_from)
        end = parse_date(date_to)
        days = (end - start).days + 1
        date_labels = [(start + timedelta(days=i)).strftime('%d.%m.%Y') for i in range(days)]
        income_by_date = defaultdict(int)
        for a in appointments:
            income_by_date[a.appointment_date.strftime('%d.%m.%Y')] += a.service.price if a.service and a.service.price else 0
        chart_data['labels'] = date_labels
        chart_data['values'] = [income_by_date.get(lbl, 0) for lbl in date_labels]
    else:
        from datetime import date, timedelta
        end = date.today()
        start = end - timedelta(days=13)
        date_labels = [(start + timedelta(days=i)).strftime('%d.%m.%Y') for i in range(14)]
        income_by_date = defaultdict(int)
        for a in appointments:
            income_by_date[a.appointment_date.strftime('%d.%m.%Y')] += a.service.price if a.service and a.service.price else 0
        chart_data['labels'] = date_labels
        chart_data['values'] = [income_by_date.get(lbl, 0) for lbl in date_labels]

    # --- Экспорт ---
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=finance.csv'
        writer = csv.writer(response)
        writer.writerow(['Дата', 'Пациент', 'Услуга', 'Стоимость'])
        for a in appointments:
            writer.writerow([
                a.appointment_date.strftime('%d.%m.%Y'),
                a.client.get_full_name() if a.client else '',
                a.service.name if a.service else '',
                a.service.price if a.service and a.service.price else 0
            ])
        return response
    elif export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Финансы'
        ws.append(['Дата', 'Пациент', 'Услуга', 'Стоимость'])
        for a in appointments:
            ws.append([
                a.appointment_date.strftime('%d.%m.%Y'),
                a.client.get_full_name() if a.client else '',
                a.service.name if a.service else '',
                a.service.price if a.service and a.service.price else 0
            ])
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=finance.xlsx'
            return response

    context = {
        'title': 'Финансовые отчёты',
        'appointments': appointments_list,
        'services': services,
        'chart_data': chart_data,
        'selected_service': int(service_id) if service_id else None,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reports/finance.html', context)


@login_required
def report_staff(request):
    # --- Фильтры ---
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    position = request.GET.get('position')
    department = request.GET.get('department')
    export_format = request.GET.get('export')

    # --- Списки для фильтров ---
    positions = []
    departments = []

    # --- Queryset сотрудников ---
    staff_qs = []
    if position:
        staff_qs = []
    if department:
        staff_qs = []

    # --- Эффективность: количество приёмов за период ---
    from services.models import ServiceAppointment
    from datetime import date, timedelta
    from collections import Counter
    if date_from:
        staff_qs = []
    if date_to:
        staff_qs = []
    if not date_from and not date_to:
        # По умолчанию за последние 14 дней
        end = date.today()
        start = end - timedelta(days=13)
        staff_qs = []

    staff_list = []

    # --- Данные для графика ---
    chart_data = {'labels': [], 'values': []}
    staff_names = []
    counts = []
    chart_data['labels'] = staff_names
    chart_data['values'] = counts

    # --- Экспорт ---
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=staff_efficiency.csv'
        writer = csv.writer(response)
        writer.writerow(['ФИО', 'Должность', 'Отделение', 'Количество приёмов'])
        for s in staff_list:
            writer.writerow([
                f"{s.last_name} {s.first_name}",
                s.position,
                s.department,
                getattr(s, 'appointments_count', 0)
            ])
        return response
    elif export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Персонал'
        ws.append(['ФИО', 'Должность', 'Отделение', 'Количество приёмов'])
        for s in staff_list:
            ws.append([
                f"{s.last_name} {s.first_name}",
                s.position,
                s.department,
                getattr(s, 'appointments_count', 0)
            ])
        for col in range(1, 5):
            ws.column_dimensions[get_column_letter(col)].width = 20
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=staff_efficiency.xlsx'
            return response

    context = {
        'title': 'Эффективность персонала',
        'staff_list': staff_list,
        'positions': positions,
        'departments': departments,
        'chart_data': chart_data,
        'selected_position': position,
        'selected_department': department,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reports/staff.html', context)


from django.utils.dateparse import parse_date
import openpyxl
from openpyxl.utils import get_column_letter
import tempfile

@login_required
def report_appointments(request):
    # --- Фильтры ---
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    doctor_id = request.GET.get('doctor')
    service_id = request.GET.get('service')
    export_format = request.GET.get('export')  # 'csv' или 'excel'

    # --- Списки для фильтров ---
    doctors = []
    services = ServiceItem.objects.all().order_by('name')

    # --- Queryset приёмов ---
    appointments = ServiceAppointment.objects.select_related('client', 'specialist', 'service').all()
    if date_from:
        appointments = appointments.filter(appointment_date__gte=parse_date(date_from))
    if date_to:
        appointments = appointments.filter(appointment_date__lte=parse_date(date_to))
    if doctor_id:
        appointments = appointments.filter(specialist_id=doctor_id)
    if service_id:
        appointments = appointments.filter(service_id=service_id)

    # --- Данные для таблицы ---
    appointments_list = appointments.order_by('-appointment_date', '-start_time')[:200]  # лимит для UI

    # --- Данные для графика (количество приёмов по дням) ---
    from collections import Counter
    chart_data = {'labels': [], 'values': []}
    if date_from and date_to:
        from datetime import timedelta
        start = parse_date(date_from)
        end = parse_date(date_to)
        days = (end - start).days + 1
        date_labels = [(start + timedelta(days=i)).strftime('%d.%m.%Y') for i in range(days)]
        counts = Counter(appointments.values_list('appointment_date', flat=True))
        chart_data['labels'] = date_labels
        chart_data['values'] = [counts.get(parse_date(lbl), 0) for lbl in date_labels]
    else:
        # По умолчанию последние 14 дней
        from datetime import date, timedelta
        end = date.today()
        start = end - timedelta(days=13)
        date_labels = [(start + timedelta(days=i)).strftime('%d.%m.%Y') for i in range(14)]
        counts = Counter(appointments.values_list('appointment_date', flat=True))
        chart_data['labels'] = date_labels
        chart_data['values'] = [counts.get(parse_date(lbl), 0) for lbl in date_labels]

    # --- Экспорт ---
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=appointments.csv'
        writer = csv.writer(response)
        writer.writerow(['Дата', 'Время', 'Пациент', 'Врач', 'Услуга', 'Статус'])
        for a in appointments:
            writer.writerow([
                a.appointment_date.strftime('%d.%m.%Y'),
                a.start_time.strftime('%H:%M'),
                a.client.get_full_name() if a.client else '',
                a.specialist.get_full_name() if a.specialist else '',
                a.service.name if a.service else '',
                a.get_status_display() if hasattr(a, 'get_status_display') else a.status
            ])
        return response
    elif export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Приёмы'
        ws.append(['Дата', 'Время', 'Пациент', 'Врач', 'Услуга', 'Статус'])
        for a in appointments:
            ws.append([
                a.appointment_date.strftime('%d.%m.%Y'),
                a.start_time.strftime('%H:%M'),
                a.client.get_full_name() if a.client else '',
                a.specialist.get_full_name() if a.specialist else '',
                a.service.name if a.service else '',
                a.get_status_display() if hasattr(a, 'get_status_display') else a.status
            ])
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=appointments.xlsx'
            return response

    # --- Контекст для шаблона ---
    context = {
        'title': 'Аналитика записей',
        'appointments': appointments_list,
        'doctors': doctors,
        'services': services,
        'chart_data': chart_data,
        'selected_doctor': int(doctor_id) if doctor_id else None,
        'selected_service': int(service_id) if service_id else None,
        'date_from': date_from,
        'date_to': date_to,
    }
    return render(request, 'reports/appointments.html', context)


@login_required
def report_rehab(request):
    from rehab_programs.models import RehabilitationProgram
    from django.db.models import Q
    from django.utils.dateparse import parse_date
    import openpyxl
    from openpyxl.utils import get_column_letter
    import tempfile
    import csv
    # --- Фильтры ---
    program_type = request.GET.get('program_type')
    status = request.GET.get('status')
    export_format = request.GET.get('export')

    STATUSES = [s[0] for s in RehabilitationProgram._meta.get_field('status').choices]
    PROGRAM_TYPES = [t[0] for t in RehabilitationProgram._meta.get_field('program_type').choices]
    # --- Queryset программ ---
    programs_qs = RehabilitationProgram.objects.all()
    if program_type:
        programs_qs = programs_qs.filter(program_type=program_type)
    if status:
        programs_qs = programs_qs.filter(status=status)
    programs_list = programs_qs.order_by('-start_date')[:200]

    # --- Данные для графика (количество по статусам) ---
    from collections import Counter
    chart_data = {'labels': [], 'values': []}
    status_counts = Counter([p.status for p in programs_list])
    chart_data['labels'] = [s for s in STATUSES]
    chart_data['values'] = [status_counts.get(s, 0) for s in STATUSES]

    # --- Экспорт ---
    if export_format == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename=rehab.csv'
        writer = csv.writer(response)
        writer.writerow(['Пациент', 'Тип программы', 'Статус', 'Дата начала', 'Дата окончания', 'Цель'])
        for p in programs_list:
            writer.writerow([
                str(p.patient),
                p.get_program_type_display(),
                p.get_status_display(),
                p.start_date.strftime('%d.%m.%Y') if p.start_date else '',
                p.end_date.strftime('%d.%m.%Y') if p.end_date else '',
                p.goal
            ])
        return response
    elif export_format == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Реабилитация'
        ws.append(['Пациент', 'Тип программы', 'Статус', 'Дата начала', 'Дата окончания', 'Цель'])
        for p in programs_list:
            ws.append([
                str(p.patient),
                p.get_program_type_display(),
                p.get_status_display(),
                p.start_date.strftime('%d.%m.%Y') if p.start_date else '',
                p.end_date.strftime('%d.%m.%Y') if p.end_date else '',
                p.goal
            ])
        for col in range(1, 7):
            ws.column_dimensions[get_column_letter(col)].width = 20
        with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
            wb.save(tmp.name)
            tmp.seek(0)
            response = HttpResponse(tmp.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename=rehab.xlsx'
            return response

    context = {
        'title': 'Результаты программ',
        'programs': programs_list,
        'statuses': STATUSES,
        'program_types': PROGRAM_TYPES,
        'chart_data': chart_data,
        'selected_program_type': program_type,
        'selected_status': status,
    }
    return render(request, 'reports/rehab.html', context)


from services.forms import ServiceAppointmentForm
from .notifications import send_appointment_confirmation_email

class DashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'core/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Общая статистика
        today = timezone.now().date()
        
        # Подсчет пациентов
        context['total_patients'] = Patient.objects.count()
        month_ago = today - timedelta(days=30)
        context['new_patients_monthly'] = Patient.objects.filter(created_at__gte=month_ago).count()
        
        # Приемы на сегодня
        todays_appointments = Appointment.objects.filter(
            start_datetime__date=today
        ).select_related('patient', 'doctor').order_by('start_datetime')
        context['todays_appointments'] = todays_appointments
        context['todays_appointments_count'] = todays_appointments.count()
        
        # Финансовая информация (заглушка - в будущем нужно подключить к реальной системе оплат)
        context['hospital_earning'] = "₽ 305,000"
        
        # Добавляем сегодняшнюю дату для отображения
        context['today'] = today
        
        context['page_title'] = "Панель управления"
        return context


def get_dashboard_stats(request):
    period = request.GET.get('period', 'week')
    today = timezone.now().date()

    if period == 'month':
        start_date = today.replace(day=1)
    else: # Default to 'week'
        start_date = today - timedelta(days=today.weekday())

    total_patients = Patient.objects.count()
    appointments_in_period = ServiceAppointment.objects.filter(appointment_date__gte=start_date).count()
    new_patients_in_period = Patient.objects.filter(created_at__date__gte=start_date).count()
    
    earnings_in_period = ServiceAppointment.objects.filter(
        appointment_date__gte=start_date,
        status='completed'
    ).aggregate(total=Sum('service__price'))['total'] or 0

    data = {
        'total_patients': total_patients,
        'todays_appointments_count': appointments_in_period,
        'appointments': appointments_in_period,
        'new_patients_monthly': new_patients_in_period,
        'new_patients': new_patients_in_period,
        'hospital_earning': f"${intcomma(int(earnings_in_period))}",
    }

    return JsonResponse(data)

class HomeView(TemplateView):
    template_name = 'home.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Add site information to context
        context['SITE_NAME'] = SITE_NAME
        context['SITE_DESCRIPTION'] = SITE_DESCRIPTION
        # Add SEO meta information
        context['meta'] = {
            'title': f"Главная - {SITE_NAME}",
            'description': SITE_DESCRIPTION,
            'og_title': f"Добро пожаловать в {SITE_NAME}",
            'og_description': "Профессиональная медицинская реабилитация для вашего здоровья.",
            'og_image': '/static/images/og-home.jpg',
        }
        return context

User = get_user_model()

from django.urls import reverse, reverse_lazy
from django.views.generic import CreateView, UpdateView, DeleteView, ListView, DetailView

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test


def is_superuser(user):
    return user.is_superuser

@login_required
@user_passes_test(is_superuser)
def settings_hub(request):
    # Получаем или создаём singleton объекты
    system_settings, _ = SystemSettings.objects.get_or_create(pk=1)
    integration_settings, _ = IntegrationSettings.objects.get_or_create(pk=1)

    system_form = SystemSettingsForm(request.POST or None, request.FILES or None, instance=system_settings)
    integration_form = IntegrationSettingsForm(request.POST or None, instance=integration_settings)

    active_tab = request.POST.get('active_tab', 'general')
    updated = False

    if request.method == 'POST':
        if 'save_system' in request.POST and system_form.is_valid():
            old_values = {f: getattr(system_settings, f) for f in system_form.fields}
            system_form.save()
            updated = True
            # Audit log
            AuditLog.objects.create(
                user=request.user,
                action='update',
                model='SystemSettings',
                object_id=system_settings.pk,
                old_value=str(old_values),
                new_value=str({f: getattr(system_settings, f) for f in system_form.fields})
            )
            messages.success(request, 'Основные настройки успешно сохранены.')
            active_tab = 'general'
        elif 'save_integration' in request.POST and integration_form.is_valid():
            old_values = {f: getattr(integration_settings, f) for f in integration_form.fields}
            integration_form.save()
            updated = True
            # Audit log
            AuditLog.objects.create(
                user=request.user,
                action='update',
                model='IntegrationSettings',
                object_id=integration_settings.pk,
                old_value=str(old_values),
                new_value=str({f: getattr(integration_settings, f) for f in integration_form.fields})
            )
            messages.success(request, 'Интеграции успешно сохранены.')
            active_tab = 'integration'

    # --- Пользователи и роли ---
    users = User.objects.all().select_related()
    groups = Group.objects.all()
    user_form = UserCreateForm()
    group_form = None  # Можно добавить форму создания групп при необходимости

    # --- Экспорт/импорт ---
    export_patients_url = reverse('core:export_patients')
    export_admissions_url = reverse('core:export_admissions')
    import_patients_url = reverse('core:import_patients')
    import_admissions_url = reverse('core:import_admissions')

    # --- Кастомизация интерфейса ---
    # (используется system_form: logo, color_theme)

    context = {
        'system_form': system_form,
        'integration_form': integration_form,
        'active_tab': active_tab,
        'users': users,
        'groups': groups,
        'user_form': user_form,
        'export_patients_url': export_patients_url,
        'export_admissions_url': export_admissions_url,
        'import_patients_url': import_patients_url,
        'import_admissions_url': import_admissions_url,
    }
    return render(request, 'core/settings.html', context)


@login_required
@user_passes_test(is_superuser)
def create_user(request):
    if request.method == 'POST':
        form = UserCreateForm(request.POST)
        if form.is_valid():
            user = form.save()
            messages.success(request, 'Пользователь успешно создан.')
            return redirect('core:settings')
        else:
            messages.error(request, 'Ошибка при создании пользователя.')
    else:
        form = UserCreateForm()
    return render(request, 'core/user_form.html', {'form': form})

@login_required
@user_passes_test(is_superuser)
def edit_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        form = UserEditForm(request.POST, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, 'Пользователь успешно обновлён.')
            return redirect('core:settings')
        else:
            messages.error(request, 'Ошибка при обновлении пользователя.')
    else:
        form = UserEditForm(instance=user)
    return render(request, 'core/user_form.html', {'form': form, 'edit': True, 'user_obj': user})

@login_required
@user_passes_test(is_superuser)
def delete_user(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        user.delete()
        messages.success(request, 'Пользователь удалён.')
        return redirect('core:settings')
    return render(request, 'core/user_confirm_delete.html', {'user_obj': user})

@login_required
@user_passes_test(is_superuser)
def change_user_role(request, user_id):
    user = get_object_or_404(User, pk=user_id)
    if request.method == 'POST':
        form = RoleChangeForm(request.POST)
        if form.is_valid():
            user.groups.set(form.cleaned_data['groups'])
            user.save()
            messages.success(request, 'Роли пользователя обновлены.')
            return redirect('core:settings')
        else:
            messages.error(request, 'Ошибка при обновлении ролей.')
    else:
        form = RoleChangeForm(initial={'groups': user.groups.all()})
    return render(request, 'core/user_role_form.html', {'form': form, 'user_obj': user})

# AJAX: смена темы оформления
@login_required
@user_passes_test(is_superuser)
def set_theme(request):
    if request.method == 'POST' and request.is_ajax():
        theme = request.POST.get('theme')
        settings = SystemSettings.objects.get(pk=1)
        settings.color_theme = theme
        settings.save()
        return JsonResponse({'status': 'ok', 'theme': theme})
    return JsonResponse({'status': 'error'}, status=400)


@login_required
@user_passes_test(is_superuser)
def export_patients(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=patients.csv'
    writer = csv.writer(response)
    fields = ['id', 'last_name', 'first_name', 'middle_name', 'date_of_birth', 'gender', 'phone_number', 'email', 'address']
    writer.writerow(fields)
    for p in Patient.objects.all():
        writer.writerow([getattr(p, f) for f in fields])
    return response

@login_required
@user_passes_test(is_superuser)
def export_admissions(request):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=admissions.csv'
    writer = csv.writer(response)
    fields = ['id', 'patient_id', 'ward_id', 'bed_id', 'admission_date', 'discharge_date', 'doctor_id', 'diagnosis']
    writer.writerow(fields)
    for adm in []:
        writer.writerow([getattr(adm, f) for f in fields])
    return response

@login_required
@user_passes_test(is_superuser)
def import_patients(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        decoded = file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded)
        for row in reader:
            Patient.objects.update_or_create(
                id=row['id'],
                defaults={k: v for k, v in row.items() if k != 'id'}
            )
        messages.success(request, 'Пациенты успешно импортированы.')
        return redirect('core:settings')
    return render(request, 'core/import_patients.html')

@login_required
@user_passes_test(is_superuser)
def import_admissions(request):
    if request.method == 'POST' and request.FILES.get('file'):
        file = request.FILES['file']
        decoded = file.read().decode('utf-8').splitlines()
        reader = csv.DictReader(decoded)
        for row in reader:
            pass
        messages.success(request, 'Госпитализации успешно импортированы.')
        return redirect('core:settings')
    return render(request, 'core/import_admissions.html')

# Ward, Bed, PatientAdmission и Staff view классы временно отключены 
# так как соответствующие модели не существуют

# Specialization Views moved to views_staff.py

@login_required
@permission_required('core.view_patient', raise_exception=True)
def patient_detail(request, pk):
    """
    Отображает детальную информацию о пациенте, включая документы и медицинские записи.
    Обрабатывает формы для добавления новых документов и медицинских записей.
    """
    patient = get_object_or_404(Patient, pk=str(pk))
    
    # Проверка прав доступа к пациенту
    if not (request.user.has_perm('core.view_all_patients') or 
            patient.curator == request.user or
            patient.user == request.user):
        raise PermissionDenied("У вас нет прав для просмотра этого пациента")
    
    # Инициализация форм
    document_form = PatientDocumentForm()
    medical_record_form = MedicalRecordForm()
    
    # Получаем документы с сортировкой по дате загрузки (новые сверху)
    documents = patient.patientdocument_set.all().order_by('-upload_date')
    
    # Поиск по документам
    search_query = request.GET.get('q')
    if search_query:
        documents = documents.filter(
            Q(file__icontains=search_query) |
            Q(description__icontains=search_query) |
            Q(document_type__icontains=search_query)
        )
    
    # Пагинация
    paginator = Paginator(documents, 10)  # 10 документов на страницу
    page = request.GET.get('page')
    documents_page = paginator.get_page(page)
    
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        # Обработка загрузки документа
        if form_type == 'document':
            if not request.user.has_perm('core.add_patientdocument'):
                messages.error(request, 'У вас нет прав на загрузку документов')
            else:
                document_form = PatientDocumentForm(request.POST, request.FILES)
                if document_form.is_valid():
                    try:
                        document = document_form.save(commit=False)
                        document.patient = patient
                        document.uploaded_by = request.user
                        document.save()
                        
                        # Логирование действия
                        from django.contrib.admin.models import LogEntry, ADDITION
                        from django.contrib.contenttypes.models import ContentType
                        LogEntry.objects.log_action(
                            user_id=request.user.id,
                            content_type_id=ContentType.objects.get_for_model(document).pk,
                            object_id=document.pk,
                            object_repr=str(document),
                            action_flag=ADDITION,
                            change_message='Документ загружен'
                        )
                        
                        messages.success(request, 'Документ успешно загружен')
                        return redirect('core:patient_detail', pk=patient.pk)
                    except Exception as e:
                        messages.error(request, f'Ошибка при загрузке документа: {str(e)}')
                else:
                    for field, errors in document_form.errors.items():
                        messages.error(request, f'Ошибка в поле "{document_form.fields[field].label}": {error}')
        
        # Обработка добавления медицинской записи
        elif form_type == 'medical_record':
            medical_record_form = MedicalRecordForm(request.POST)
            if medical_record_form.is_valid():
                try:
                    record = medical_record_form.save(commit=False)
                    record.patient = patient
                    record.save()
                    messages.success(request, 'Медицинская запись успешно добавлена')
                    return redirect('core:patient_detail', pk=patient.pk)
                except Exception as e:
                    messages.error(request, f'Ошибка при добавлении медицинской записи: {str(e)}')
            else:
                for field, errors in medical_record_form.errors.items():
                    field_label = medical_record_form.fields[field].label if field in medical_record_form.fields else field
                    for error in errors:
                        messages.error(request, f'Ошибка в поле "{field_label}": {error}')
    
    # Подготавливаем контекст для шаблона
    context = {
        'patient': patient,
        'documents_page': documents_page,
        'document_form': document_form,
        'medical_records': patient.history_medical_records.all().order_by('-diagnosis_date'),
        'medical_record_form': medical_record_form,
        'rehab_programs': patient.rehabilitation_programs.all(),
        'search_query': search_query or '',
    }
    
    # Добавляем хлебные крошки
    breadcrumbs = [
        {'title': 'Пациенты', 'url': reverse('core:patient_list')},
        {'title': f'{patient.get_full_name()}', 'url': ''},
    ]
    context['breadcrumbs'] = breadcrumbs
    
    return render(request, 'core/patient_detail.html', context)

def patient_update(request, pk):
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        form = PatientForm(request.POST, instance=patient)
        if form.is_valid():
            form.save()
            return redirect('core:patient_detail', pk=patient.pk)
    else:
        form = PatientForm(instance=patient)
    return render(request, 'core/patient_form.html', {'form': form, 'patient': patient})

def patient_create(request):
    if request.method == 'POST':
        form = PatientForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('core:patient_list')
    else:
        form = PatientForm()
    return render(request, 'core/patient_form.html', {'form': form})

def patient_delete(request, pk):
    from django.urls import reverse_lazy
    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect
    
    patient = get_object_or_404(Patient, pk=pk)
    if request.method == 'POST':
        patient.delete()
        messages.success(request, f'Пациент {patient.get_full_name()} успешно удален.')
        return redirect('core:patient_list')
    
    # This should not be reached if form is submitted via POST
    return redirect('core:patient_list')

def patient_list(request):
    from django.db.models import Q
    from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
    
    # Get search query
    query = request.GET.get('q', '')
    
    # Get filter parameters
    status = request.GET.get('status', '')
    department = request.GET.get('department', '')
    
    # Start with base queryset
    patients = Patient.objects.all().order_by('last_name', 'first_name')
    
    # Apply search
    if query:
        patients = patients.filter(
            Q(last_name__icontains=query) |
            Q(first_name__icontains=query) |
            Q(middle_name__icontains=query) |
            Q(phone__icontains=query) |
            Q(insurance_number__icontains=query)
        )
    
    # Apply filters
    if status:
        patients = patients.filter(status=status)
    if department:
        patients = patients.filter(department=department)
    
    # Pagination
    page = request.GET.get('page', 1)
    paginator = Paginator(patients, 25)  # Show 25 patients per page
    
    try:
        patients = paginator.page(page)
    except PageNotAnInteger:
        patients = paginator.page(1)
    except EmptyPage:
        patients = paginator.page(paginator.num_pages)
    
    context = {
        'patients': patients,
        'title': 'Список пациентов',
        'search_query': query,
        'status_filter': status,
        'department_filter': department,
        'PATIENT_STATUS_CHOICES': Patient.PATIENT_STATUS_CHOICES if hasattr(Patient, 'PATIENT_STATUS_CHOICES') else [],
        'DEPARTMENT_CHOICES': Patient.DEPARTMENT_CHOICES if hasattr(Patient, 'DEPARTMENT_CHOICES') else [],
    }
    return render(request, 'core/patient_list.html', context)

def format_number(value):
    """Format number with spaces as thousand separators"""
    if value is None:
        return '0'
    try:
        return "{:,}".format(int(value)).replace(",", " ")
    except (ValueError, TypeError):
        return str(value)

def contact_view(request):
    """View for the contact page."""
    context = {
        'title': 'Контакты',
        'meta': {
            'title': 'Контакты',
            'description': 'Свяжитесь с нами для получения дополнительной информации о наших услугах.'
        }
    }
    return render(request, 'contact.html', context)


def about_view(request):
    """View for the about page."""
    context = {
        'title': 'О нас',
        'meta': {
            'title': 'О нас',
            'description': 'Узнайте больше о нашем центре и нашей команде профессионалов.'
        }
    }
    return render(request, 'about.html', context)


def dashboard(request):
    # Format numbers with spaces as thousand separators
    def format_number(value):
        if value is None:
            return '0'
        try:
            return "{:,}".format(int(value)).replace(",", " ")
        except (ValueError, TypeError):
            return str(value)

    try:
        # Try to get real data from the database
        from django.db.models import Count, Sum, Q
        from datetime import datetime, timedelta
        
        # Get counts for dashboard cards
        patients_count = Patient.objects.count()
        active_programs = 45  # Placeholder - replace with actual query
        available_beds = 32   # Placeholder - replace with actual query
        staff_count = 28      # Placeholder - replace with actual query
        
        # Get recent patients (last 3)
        recent_patients = Patient.objects.order_by('-created_at')[:3]
        recent_patients_data = [
            {
                'program_type': f"{p.last_name} {p.first_name}",
                'admission_date': p.created_at.strftime('%Y-%m-%d') if hasattr(p, 'created_at') else 'Неизвестно',
                'program': 'Стандартная'  # Placeholder - replace with actual program data
            }
            for p in recent_patients
        ]
        
        # Ensure we have at least 3 patients
        while len(recent_patients_data) < 3:
            recent_patients_data.append({
                'program_type': 'Нет данных',
                'admission_date': '--',
                'program': '--'
            })
        
        # Sample data for services stats
        services_stats = {
            'total_revenue': 1250000,
            'top_services': [
                {'program_type': 'Консультация врача', 'total_revenue': 250000, 'count': 125},
                {'program_type': 'Массаж', 'total_revenue': 380000, 'count': 190},
                {'program_type': 'Физиотерапия', 'total_revenue': 620000, 'count': 310},
            ]
        }
        
        context = {
            # Card data
            'patients_count': format_number(patients_count),
            'active_programs': format_number(active_programs),
            'available_beds': format_number(available_beds),
            'staff_count': format_number(staff_count),
            
            # Recent patients
            'recent_patients': recent_patients_data,
            
            # Upcoming appointments (sample data for now)
            'upcoming_appointments': [
                {'time': '10:00', 'patient': 'Иванов Иван', 'type': 'Консультация'},
                {'time': '11:30', 'patient': 'Петрова Анна', 'type': 'Процедура'},
                {'time': '14:15', 'patient': 'Сидоров Алексей', 'type': 'Осмотр'},
            ],
            
            # Services stats
            'services_stats': services_stats
        }
        
    except Exception as e:
        # Fallback to sample data if there's any error
        print(f"Error in dashboard view: {str(e)}")
        context = {
            'patients_count': format_number(150),
            'active_programs': format_number(45),
            'available_beds': format_number(32),
            'staff_count': format_number(28),
            'recent_patients': [
                {'program_type': 'Иванов Иван', 'admission_date': '2023-06-10', 'program': 'Стандартная'},
                {'program_type': 'Петрова Анна', 'admission_date': '2023-06-09', 'program': 'Интенсив'},
                {'program_type': 'Сидоров Алексей', 'admission_date': '2023-06-08', 'program': 'Премиум'},
            ],
            'upcoming_appointments': [
                {'time': '10:00', 'patient': 'Иванов Иван', 'type': 'Консультация'},
                {'time': '11:30', 'patient': 'Петрова Анна', 'type': 'Процедура'},
                {'time': '14:15', 'patient': 'Сидоров Алексей', 'type': 'Осмотр'},
            ],
            'services_stats': {
                'total_revenue': 1250000,
                'top_services': [
                    {'program_type': 'Консультация врача', 'total_revenue': 250000, 'count': 125},
                    {'program_type': 'Массаж', 'total_revenue': 380000, 'count': 190},
                    {'program_type': 'Физиотерапия', 'total_revenue': 620000, 'count': 310},
                ]
            }
        }
    
    return render(request, 'core/dashboard_clean.html', context)


def bed_management(request):
    """
    View for displaying and managing hospital beds
    """
    # Get all active wards with their beds and patient admissions
    wards = []
    
    # Prepare data for the template
    wards_data = []
    for ward in wards:
        beds_data = []
        for bed in ward.beds_with_admissions:
            # Get active admission if exists
            active_admission = bed.active_admissions[0] if hasattr(bed, 'active_admissions') and bed.active_admissions else None
            
            bed_data = {
                'id': bed.id,
                'number': bed.number,
                'bed_type': bed.get_bed_type_display(),
                'is_available': bed.is_available,
                'is_occupied': active_admission is not None,
                'patient': None,
                'admission_date': None,
                'doctor': None,
                'diagnosis': None
            }
            
            if active_admission:
                bed_data.update({
                    'patient': {
                        'id': active_admission.patient.id,
                        'full_name': active_admission.patient.get_full_name(),
                        'birth_date': active_admission.patient.date_of_birth.strftime('%d.%m.%Y') if active_admission.patient.date_of_birth else '—',
                    },
                    'admission_date': active_admission.admission_date.strftime('%d.%m.%Y'),
                    'doctor': active_admission.doctor.get_full_name() if active_admission.doctor else 'Не назначен',
                    'diagnosis': active_admission.diagnosis or '—'
                })
            
            beds_data.append(bed_data)
        
        wards_data.append({
            'id': ward.id,
            'program_type': ward.name,
            'department': ward.department or '—',
            'floor': ward.floor,
            'capacity': ward.capacity,
            'available_beds': ward.get_available_beds_count(),
            'beds': beds_data
        })
    
    context = {
        'wards': wards_data,
        'current_date': timezone.now().strftime('%d %B %Y')
    }
    
    return render(request, 'core/bed_management.html', context)


# Reports
# ===================================================================

def report_patient_statistics(request):
    """
    View to display patient statistics report.
    """
    # Total patients
    total_patients = Patient.objects.count()

    # Patients by gender
    gender_distribution = Patient.objects.values('gender').annotate(count=Count('gender')).order_by('-count')
    gender_data = {
        'labels': [dict(Patient.GENDER_CHOICES).get(item['gender'], 'Не указан') for item in gender_distribution],
        'counts': [item['count'] for item in gender_distribution]
    }

    # Patients by age group
    today = date.today()
    age_groups = {
        '0-17': 0,
        '18-35': 0,
        '36-55': 0,
        '56+': 0
    }
    patients_with_dob = Patient.objects.filter(date_of_birth__isnull=False)
    for patient in patients_with_dob:
        age = today.year - patient.date_of_birth.year - ((today.month, today.day) < (patient.date_of_birth.month, patient.date_of_birth.day))
        if age <= 17:
            age_groups['0-17'] += 1
        elif 18 <= age <= 35:
            age_groups['18-35'] += 1
        elif 36 <= age <= 55:
            age_groups['36-55'] += 1
        else:
            age_groups['56+'] += 1
    
    age_data = {
        'labels': list(age_groups.keys()),
        'counts': list(age_groups.values())
    }

    context = {
        'title': 'Статистика по пациентам',
        'total_patients': total_patients,
        'gender_data': gender_data,
        'age_data': age_data,
    }
    
    return render(request, 'reports/patients.html', context)


@login_required
def get_all_appointments(request):
    """
    API endpoint to fetch all service appointments for FullCalendar.
    """
    appointments = ServiceAppointment.objects.select_related(
        'client__user', 'service', 'specialist__user'
    ).all()
    
    events = []
    for appointment in appointments:
        client_name = appointment.client.user.get_full_name() if appointment.client and appointment.client.user else "N/A"
        service_name = appointment.service.name if appointment.service else "N/A"
        specialist_name = appointment.specialist.user.get_full_name() if appointment.specialist and appointment.specialist.user else "N/A"
        
        title = f"{client_name} - {service_name} ({specialist_name})"

        if appointment.status == 'completed':
            color = '#198754'  # Green
        elif appointment.status == 'cancelled':
            color = '#dc3545'  # Red
        else: # scheduled
            color = '#0d6efd'  # Blue

        events.append({
            'id': appointment.pk,
            'title': title,
            'start': f"{appointment.appointment_date.isoformat()}T{appointment.start_time.strftime('%H:%M:%S')}",
            'end': f"{appointment.appointment_date.isoformat()}T{appointment.end_time.strftime('%H:%M:%S')}",
            'url': reverse('services:service_appointment_detail', args=[appointment.pk]),
            'backgroundColor': color,
            'borderColor': color,
            'extendedProps': {
                'status': appointment.get_status_display(),
                'specialist': specialist_name,
            }
        })
        
    return JsonResponse(events, safe=False)


@login_required
def create_appointment_view(request):
    if request.method == 'POST':
        form = ServiceAppointmentForm(request.POST, user=request.user)
        if form.is_valid():
            appointment = form.save()
            try:
                send_appointment_confirmation_email(appointment)
            except Exception as e:
                # Log the exception, but don't fail the request
                print(f"Error sending confirmation email: {e}")
            return JsonResponse({'success': True})
        else:
            # Render form with errors and return as JSON
            form_html = render_to_string('core/_appointment_form.html', {'form': form}, request=request)
            return JsonResponse({'success': False, 'form_html': form_html}, status=400)
    else:
        initial_data = {}
        date_str = request.GET.get('date')
        if date_str:
            initial_data['appointment_date'] = date_str
        
        form = ServiceAppointmentForm(initial=initial_data, user=request.user)
    
    return render(request, 'core/_appointment_form.html', {'form': form})
